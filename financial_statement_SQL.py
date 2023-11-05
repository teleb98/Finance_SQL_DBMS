# "C:\test files\Basics of stock data\modified" 폴더의 기업들의 재무제표를 담은 엑셀 파일들을 읽고
# 하나의 데이터베이스에 테이블로 저장하는 코드

import sqlite3
import glob
import os
import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm

class FinancialStatementDatabase:
    def __init__(self, db_path, folder_path):
        self.db_path = db_path
        self.folder_path = folder_path
        self.conn = sqlite3.connect(self.db_path)
        self.cursor = self.conn.cursor()

    def read_excel_files(self):
        # 경로에서 모든 엑셀 파일을 찾습니다.
        excel_files = glob.glob(os.path.join(self.folder_path, '*.xlsx'))
        # 진행 상황을 표시하기 위해 tqdm을 사용합니다.
        for file in tqdm(excel_files):
            self.process_file(file)

    def process_file(self, file_path):
        # 파일명에서 기업명을 추출합니다. 파일명이 '기업명_기타정보.xlsx' 형태라고 가정합니다.
        company_name = os.path.basename(file_path).split()[0]
        workbook = load_workbook(filename=file_path, data_only=True)
        # 모든 시트를 순회합니다.
        for sheet_name in workbook.sheetnames:
            self.process_sheet(workbook[sheet_name], company_name, sheet_name)

    def process_sheet(self, sheet, company_name, sheet_name):
        # 판다스를 사용하여 데이터프레임으로 변환합니다.
        data = sheet.values
        columns = next(data)[0:]  # 첫 번째 행을 컬럼으로 사용합니다.
        
        # 중복된 컬럼 이름을 처리합니다.
        columns = self.handle_duplicate_columns(columns)
        
        df = pd.DataFrame(data, columns=columns)
        # 테이블 이름을 '기업명_시트명'으로 정합니다.
        table_name = f"{company_name}_{sheet_name}"
        # 데이터베이스에 저장합니다.
        df.to_sql(table_name, self.conn, if_exists='replace', index=False)
    
    def handle_duplicate_columns(self, columns):
        # 중복 컬럼 이름이 있는지 검사하고, 있다면 고유하게 만듭니다.
        new_columns, counts = [], {}
        for col in columns:
            if col in counts:
                counts[col] += 1
                new_columns.append(f"{col}_{counts[col]}")
            else:
                counts[col] = 1
                new_columns.append(col)
        return new_columns
    
    def close_connection(self):
        # 데이터베이스 연결을 안전하게 종료합니다.
        self.conn.close()

# 사용 예:
folder_path = r"C:\test files\Basics of stock data\modified"
db_path = r"C:\test files\Finance_SQL_DB_MS\DB\financial_statements.db"
database = FinancialStatementDatabase(db_path, folder_path)
database.read_excel_files()
database.close_connection()